"""Defines CDT of the continuous/untrimmed datasets."""
import numpy as np
from .skeleton_abstract import SkeletonDataset
from global_configs import DatasetProtocol, SensorJointNumber
from utils.misc import ProgressBar, get_folders_and_files, is_file_empty
from utils.processing import (preprocess_skeleton_frame,
                              causal_savitzky_golay_filter,
                              standardize_coordinate_origin_sequence)


__all__ = ['SkeletonDatasetPKUMMDv1', 'SkeletonDatasetPKUMMDv2', 'SkeletonDatasetOAD', 'SkeletonDatasetG3D',
           'SkeletonDatasetUTKinect']
np.set_printoptions(threshold=5000)
PKUMMD_INTERACT_LABELS = np.array([12, 14, 16, 18, 21, 24, 26, 27]) - 1


class SkeletonDatasetPKUMMDv1(SkeletonDataset):
    """
    Load the PKU-MMD dataset by:
        C. Liu, Y. Hu, Y. Li, S. Song, and J. Liu, "PKU-MMD: A large scale benchmark for continuous multi-modal human
        action understanding," arXiv preprint arXiv:1703.07475, 2017.

    Dataset description:
        Phase #1 contains 1076 long video sequences in 51 action categories, performed by 66 subjects in three camera
        views. It contains almost 20,000 action instances and 5.4 million frames in total. Each video lasts about 3-4
        minutes (recording ratio set to 30 FPS) and contains approximately 20 action instances. The total scale of our
        dataset is 5,312,580 frames of 3,000 minutes with 21,545 temporally localized actions. We choose 51 action
        classes in total, which are divided into two parts: 41 daily actions (drinking, waving hand, putting on the
        glassed, etc.) and 10 interaction actions (hugging, shaking hands, etc.). 66 distinct subjects are invited
        for our data collection. Each subjects takes part in 4 daily action videos and 2 interactive action
        videos.our videos only contain one part of the actions, either daily actions or interaction actions. We
        design 54 sequences and divide subjects into 9 groups, and each groups randomly choose 6 sequences to perform.

        Skeleton Files:
            For each video, there exists a skeleton file XXXX−V.skeletonXXXX−V.skeleton which contains several lines for
            frame-level skeleton data. Each line contains 3×25×23×25×2 float numbers for 3-dimensional locations of 25
            major body joints of 2 subjects.

        Label Files:
            For each video, there exists a label file named XXXX−V.labelXXXX−V.label illustrating the ground truth
            labels. Several lines will be given, each line contains 4 integers for label,start,end,confidence
            respectively. Note that confidenceconfidence is either 11 or 22 for slight and strong recommendation
            respectively.

        Split Files:
            a)	cross-view.txt: cross view split list.
            b)	cross-subject.txt: cross subject split list.
            These are the split settings in our own experiments. We split the training data into two parts for training
            and validation to tune our model for cross-view and cross-subject settings, respectively. Please note
            that you can use all the training data to train your model for the final testing evaluation.
            c)	Actions.xlsx: contains the 51 defined actions and corresponding IDs.
    """
    def __init__(self, directory: str, protocol: DatasetProtocol,
                 regression_sigma: float = 5,
                 is_translated: bool = True,
                 is_edged: bool = False,
                 is_rotated: bool = False,
                 is_filtered: bool = False):
        if not protocol:
            protocol = DatasetProtocol.CROSS_SUBJECT
        if protocol == DatasetProtocol.CROSS_SUBJECT:
            self.split_filename = 'cross-subject.txt'
        elif protocol == DatasetProtocol.CROSS_VIEW:
            self.split_filename = 'cross-view.txt'
        else:
            raise ValueError('Protocol specified is not available for this dataset.')
        super(SkeletonDatasetPKUMMDv1, self).__init__(directory, regression_sigma, 0.5,
                                                      is_translated=is_translated,
                                                      is_edged=is_edged,
                                                      is_rotated=is_rotated,
                                                      is_filtered=is_filtered,
                                                      downsample_factor=4,
                                                      forecast_t=2)

        # Load label names here instead. The sorting in parent class isn't necessary for this dataset.
        import openpyxl
        book = openpyxl.load_workbook(self.root_dir + 'Split/Actions.xlsx')
        sheet = book.active
        self._labels = []
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                self._labels.append(cell.value)
        self._labels = tuple(self._labels + ['~'])
        del sheet, book, openpyxl
        self._protocol = protocol

    def load_label_map(self):
        self.has_interaction = True
        self._joints_per_person = SensorJointNumber.KINECT_V1
        # self._joints_per_person = SensorJointNumber.KINECT_V2

        # Load train and test data file lists
        with open(self.root_dir + 'Split/' + self.split_filename, 'r') as split_file:
            split_file.readline()
            train_files = split_file.readline().rstrip().split(', ')
            split_file.readline()
            test_files = split_file.readline().rstrip().split(', ')

        # Load individual data files
        pr = ProgressBar(80, len(train_files) + len(test_files))
        data_count = 0
        for file_id in train_files:
            self._data_label_pairs.append(self._load_sequence_and_label_vector(file_id))
            self._train_indices.append(data_count)
            data_count += 1
            pr.update(data_count)

        for file_id in test_files:
            self._data_label_pairs.append(self._load_sequence_and_label_vector(file_id))
            self._test_indices.append(data_count)
            data_count += 1
            pr.update(data_count)

    def _load_sequence_and_label_vector(self, file_id):
        is_25_joint = (self._joints_per_person == SensorJointNumber.KINECT_V2)
        seq = np.array([], dtype=np.float32).reshape(0, self._joints_per_person * 3 * 2)
        edged_seq = np.array([], dtype=np.float32).reshape(0, self._joints_per_person * 3 * 2)
        with open(self.root_dir + 'Data/skeleton/' + file_id + '.txt', 'r') as skeleton_file:
            for idx, line in enumerate(skeleton_file, 0):
                frame_arr = np.array(line.rstrip().split(), dtype=np.float32) * 1000
                frame_arr_pt1 = preprocess_skeleton_frame(frame_arr[:self._joints_per_person * 3],
                                                          is_15_joint=False,
                                                          to_rotate=self.is_rotated,
                                                          to_edge=False)
                frame_arr_pt2 = preprocess_skeleton_frame(frame_arr[len(frame_arr) // 2:
                                                                    len(frame_arr) // 2 + self._joints_per_person * 3],
                                                          is_15_joint=False,
                                                          to_rotate=self.is_rotated,
                                                          to_edge=False)
                frame_arr_pt1_edged = preprocess_skeleton_frame(frame_arr[:self._joints_per_person * 3],
                                                                is_15_joint=False,
                                                                to_edge=True,
                                                                to_rotate=self.is_rotated,
                                                                is_25_joint=is_25_joint)
                frame_arr_pt2_edged = preprocess_skeleton_frame(frame_arr[len(frame_arr) // 2:
                                                                len(frame_arr) // 2 + self._joints_per_person * 3,
                                                                ],
                                                                is_15_joint=False,
                                                                to_edge=True,
                                                                to_rotate=self.is_rotated,
                                                                is_25_joint=is_25_joint)
                frame_arr = np.concatenate((frame_arr_pt1, frame_arr_pt2))
                frame_arr_edged = np.concatenate((frame_arr_pt1_edged, frame_arr_pt2_edged))
                seq = np.vstack((seq, frame_arr))
                edged_seq = np.vstack((edged_seq, frame_arr_edged))
        label_vector = 51 * np.ones(len(seq), dtype=np.int16)  # maximum means unknown (~) action
        with open(self.root_dir + 'Label/' + file_id + '.txt', 'r') as label_file:
            for idx, line in enumerate(label_file, 0):
                label_id, start_frame, end_frame = line.rstrip().split(',')[:3]
                start_frame = int(start_frame) - 1 - self.forecast_t
                if start_frame < 0:
                    start_frame = 0
                label = int(label_id) - 1
                label_vector[start_frame:int(end_frame)] = label
        if self.is_filtered:
            seq = causal_savitzky_golay_filter(seq)
            edged_seq = causal_savitzky_golay_filter(edged_seq)
        if self.is_edged and not self.is_translated:
            seq = edged_seq
        elif self.is_translated:
            seq = standardize_coordinate_origin_sequence(seq, is_25_joint=is_25_joint, use_hip=False)
            if self.is_edged:
                seq = np.concatenate((seq, edged_seq), axis=-1)
        self._update_extrema(seq)
        return seq, label_vector

    def load_data_by_idx(self, data_idx: int):
        return self._data_label_pairs[data_idx]

    # overrides
    def get_data_arrays(self, idx):
        return self.load_data_by_idx(idx)

    def __str__(self):
        return 'PKU-MMD_v1_Dataset'


class SkeletonDatasetPKUMMDv2(SkeletonDataset):
    """
    Load the PKU-MMD dataset by:
        C. Liu, Y. Hu, Y. Li, S. Song, and J. Liu, "PKU-MMD: A large scale benchmark for continuous multi-modal human
        action understanding," arXiv preprint arXiv:1703.07475, 2017.

    Dataset description:
        Phase #2 contains 2000 short video sequences in 49 action categories, performed by 13 subjects in three
        camera views. Each video lasts about 1-2 minutes (recording ratio set to 30 FPS) and contains approximately
        7 action instances.

        Skeleton Files:
            For each video, there exists a skeleton file XXXX−V.skeletonXXXX−V.skeleton which contains several lines for
            frame-level skeleton data. Each line contains 3×25×2 float numbers for 3-dimensional locations of 25
            major body joints of 2 subjects.

        Label Files:
            For each video, there exists a label file named XXXX−V.labelXXXX−V.label illustrating the ground truth
            labels. Several lines will be given, each line contains 4 integers for label,start,end,confidence
            respectively. Note that confidenceconfidence is either 11 or 22 for slight and strong recommendation
            respectively.

        Split Files:
            a)	cross-view.txt: cross view split list.
            b)	cross-subject.txt: cross subject split list.
            These are the split settings in our own experiments. We split the training data into two parts for training
            and validation to tune our model for cross-view and cross-subject settings, respectively. Please note
            that you can use all the training data to train your model for the final testing evaluation.
            c)	Actions.xlsx: contains the 51 defined actions and corresponding IDs.

        Notes:
            The following numbers for view M are missing (and therefore there are actually 338+333+338 videos in
            total): A05N04, A07N12, A08N12, A14N12, A20N01
    """
    def __init__(self, directory: str, protocol: DatasetProtocol,
                 regression_sigma: float = 5,
                 is_translated: bool = True,
                 is_edged: bool = False,
                 is_rotated: bool = False,
                 is_filtered: bool = False):
        if not protocol:
            protocol = DatasetProtocol.CROSS_SUBJECT
        if protocol == DatasetProtocol.CROSS_SUBJECT:
            self.split_filename = 'cross_subject_v2.txt'
        elif protocol == DatasetProtocol.CROSS_VIEW:
            self.split_filename = 'cross_view_v2.txt'
        else:
            raise ValueError('Protocol specified is not available for this dataset.')
        super(SkeletonDatasetPKUMMDv2, self).__init__(directory, regression_sigma, 0.5,
                                                      is_translated=is_translated,
                                                      is_edged=is_edged,
                                                      is_rotated=is_rotated,
                                                      is_filtered=is_filtered,
                                                      downsample_factor=2,
                                                      forecast_t=10)

        # Load label names here instead. The sorting in parent class isn't necessary for this dataset.
        import openpyxl
        book = openpyxl.load_workbook(self.root_dir + 'Split/Actions_v2.xlsx')
        sheet = book.active
        self._labels = []
        for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=2, max_col=2)):
            for cell in row:
                self._labels.append(cell.value)
        self._labels = tuple(self._labels + ['~'])
        del sheet, book, openpyxl
        self._protocol = protocol

    def load_label_map(self):
        self.has_interaction = True
        self._joints_per_person = SensorJointNumber.KINECT_V1
        # self._joints_per_person = SensorJointNumber.KINECT_V2

        # Load train and test data file lists
        with open(self.root_dir + 'Split/' + self.split_filename, 'r') as split_file:
            split_file.readline()
            train_files = split_file.readline().rstrip().split(', ')
            split_file.readline()
            test_files = split_file.readline().rstrip().split(', ')

        # Load individual data files
        pr = ProgressBar(80, len(train_files) + len(test_files))
        data_count = 0
        for file_id in train_files:
            self._data_label_pairs.append(self._load_sequence_and_label_vector(file_id))
            self._train_indices.append(data_count)
            data_count += 1
            pr.update(data_count)

        for file_id in test_files:
            self._data_label_pairs.append(self._load_sequence_and_label_vector(file_id))
            self._test_indices.append(data_count)
            data_count += 1
            pr.update(data_count)

    def _load_sequence_and_label_vector(self, file_id):
        is_25_joint = (self._joints_per_person == SensorJointNumber.KINECT_V2)
        seq = np.array([], dtype=np.float32).reshape(0, self._joints_per_person * 3 * 2)
        edged_seq = np.array([], dtype=np.float32).reshape(0, self._joints_per_person * 3 * 2)
        with open(self.root_dir + 'Data/skeleton/' + file_id + '.txt', 'r') as skeleton_file:
            for idx, line in enumerate(skeleton_file, 0):
                frame_arr = np.array(line.rstrip().split(), dtype=np.float32) * 1000
                frame_arr_pt1 = preprocess_skeleton_frame(frame_arr[:self._joints_per_person * 3],
                                                          is_15_joint=False,
                                                          to_edge=False,
                                                          to_rotate=self.is_rotated)
                frame_arr_pt2 = preprocess_skeleton_frame(frame_arr[len(frame_arr) // 2:
                                                                    len(frame_arr) // 2 + self._joints_per_person * 3,
                                                                    ],
                                                          is_15_joint=False,
                                                          to_edge=False,
                                                          to_rotate=self.is_rotated)
                frame_arr_pt1_edged = preprocess_skeleton_frame(frame_arr[:self._joints_per_person * 3],
                                                                is_15_joint=False,
                                                                to_edge=True,
                                                                to_rotate=self.is_rotated,
                                                                is_25_joint=is_25_joint)
                frame_arr_pt2_edged = preprocess_skeleton_frame(frame_arr[len(frame_arr) // 2:
                                                                len(frame_arr) // 2 + self._joints_per_person * 3,
                                                                ],
                                                                is_15_joint=False,
                                                                to_edge=True,
                                                                to_rotate=self.is_rotated,
                                                                is_25_joint=is_25_joint)
                frame_arr = np.concatenate((frame_arr_pt1, frame_arr_pt2))
                frame_arr_edged = np.concatenate((frame_arr_pt1_edged, frame_arr_pt2_edged))
                seq = np.vstack((seq, frame_arr))
                edged_seq = np.vstack((edged_seq, frame_arr_edged))
        label_vector = 51 * np.ones(len(seq), dtype=np.int16)  # maximum means unknown (~) action
        with open(self.root_dir + 'Label/' + file_id + '.txt', 'r') as label_file:
            for idx, line in enumerate(label_file, 0):
                label_id, start_frame, end_frame = line.rstrip().split(',')[:3]
                start_frame = int(start_frame) - 1 - self.forecast_t
                if start_frame < 0:
                    start_frame = 0
                label_vector[start_frame:int(end_frame)] = int(label_id) - 1
        if self.is_filtered:
            seq = causal_savitzky_golay_filter(seq)
            edged_seq = causal_savitzky_golay_filter(edged_seq)
        if self.is_edged and not self.is_translated:
            seq = edged_seq
        elif self.is_translated:
            seq = standardize_coordinate_origin_sequence(seq, is_25_joint=is_25_joint, use_hip=False)
            if self.is_edged:
                seq = np.concatenate((seq, edged_seq), axis=-1)
        self._update_extrema(seq)
        return seq, label_vector

    def load_data_by_idx(self, data_idx: int):
        return self._data_label_pairs[data_idx]

    # overrides
    def get_data_arrays(self, idx):
        return self.load_data_by_idx(idx)

    def __str__(self):
        return 'PKU-MMD_v2_Dataset'


class SkeletonDatasetOAD(SkeletonDataset):
    """
    Load the OAD dataset by:
        Y. Li, C. Lan, J. Xing, W. Zeng, C. Yuan, and J. Liu, "Online human action detection using joint
        classification-regression recurrent neural networks," in European Conference on Computer Vision,
        2016, pp. 203-220: Springer.

    Dataset description:
        The Online Action Detection Dataset (OAD) was captured using the Kinect V2 sensor, which collects color images,
        depth images and human skeleton joints synchronously. Our dataset includes 59 long sequences and 10 actions,
        including drinking, eating, writing, opening cup- board, washing hands, opening microwave, sweeping, gargling,
        throwing trash, and wiping.

        The folder 'data' contains the data of each sequences, including
            3. skeleton: skeleton data
            4. label: action labels. For example, the label
                drinking

                120 130

                1847 1853

                eating

                207 220
            indicates that this sequence contains 2 intervals of the action 'drinking'. The first interval starts from
            the frame indexed by 120 and ends at the frame indexed by 130.

        Note that at the beginning of some sequences, skeleton and depth files may be empty. You can simply ignore them.
        In our paper, we select 30 sequences ([1, 2, 3, 4, 7, 8, 9, 14, 15, 16, 18, 19, 20, 22, 23, 24, 25, 32, 33, 34,
        35, 37, 38, 39, 49, 50, 51, 54, 57, 58]) for training and 20 sequences ([0, 10, 13, 17, 21, 26, 27, 28, 29, 36,
        40, 41, 42, 43, 44, 45, 52, 53, 55, 56]) for testing. The remaining 9 long videos are used for the evaluation
        of the running speed.
    """
    def __init__(self, directory: str,
                 train_portion: float = 0.5,
                 regression_sigma: float = 5,
                 is_translated: bool = True,
                 is_rotated: bool = True,
                 is_filtered: bool = True,
                 is_edged: bool = False,
                 *args,
                 **kwargs):
        super(SkeletonDatasetOAD, self).__init__(directory, regression_sigma, train_portion,
                                                 is_translated=is_translated,
                                                 is_edged=is_edged,
                                                 is_rotated=is_rotated,
                                                 is_filtered=is_filtered,
                                                 forecast_t=10,
                                                 downsample_factor=1)
        # Reload label names here as the sorting in parent class isn't necessary for this dataset.
        self._labels = ('drinking', 'eating', 'writing', 'opening cupboard', 'washing hands',
                        'opening microwave oven', 'sweeping', 'gargling', 'throwing trash', 'wiping', '~')

    def load_label_map(self):
        import re
        labels = ('drinking', 'eating', 'writing', 'opening cupboard', 'washing hands', 'opening microwave oven',
                  'sweeping', 'gargling', 'throwing trash', 'wiping')
        # self._joints_per_person = SensorJointNumber.KINECT_V2
        self._joints_per_person = SensorJointNumber.KINECT_V1
        is_25_joint = (self._joints_per_person == SensorJointNumber.KINECT_V2)
        data_count = 59
        data_iter = range(data_count)
        pr = ProgressBar(80, data_count)
        self._train_indices = [1, 2, 3, 4, 7, 8, 9, 14, 15, 16, 18, 19, 20, 22, 23, 24, 25, 32, 33, 34,
                               35, 37, 38, 39, 49, 50, 51, 54, 57, 58]
        self._test_indices = [0, 10, 13, 17, 21, 26, 27, 28, 29, 36,
                              40, 41, 42, 43, 44, 45, 52, 53, 55, 56]
        data_path = self.root_dir + 'data/'
        for data_idx in data_iter:
            skeleton_dir_path = data_path + str(data_idx) + '/skeleton/'
            label_file_path = data_path + str(data_idx) + '/label/label.txt'
            _, frame_filenames = get_folders_and_files(skeleton_dir_path)
            frame_file_ids = []
            for frame_filename in frame_filenames:  # unordered
                if frame_filename.endswith('.txt'):
                    frame_file_ids.append(int(re.findall(r'\d+', frame_filename)[0]))
                else:
                    continue
            frame_file_ids = sorted(frame_file_ids)
            offset = frame_file_ids[0]  # actual start index of this sequence in the raw dataset
            seq = np.array([], dtype=np.float32).reshape(0, self._joints_per_person * 3)
            edged_seq = np.array([], dtype=np.float32).reshape(0, self._joints_per_person * 3)
            record_started = False
            for frame_id in frame_file_ids:
                frame_file_path = skeleton_dir_path + str(frame_id) + '.txt'
                if is_file_empty(frame_file_path):
                    if not record_started:
                        offset = frame_id + 1
                    continue
                else:
                    record_started = True
                with open(skeleton_dir_path + str(frame_id) + '.txt', 'r') as frame_file:
                    frame = []
                    for joint_idx in range(self._joints_per_person):
                        x, y, z = frame_file.readline().rstrip().split()
                        frame += [float(x) * 1000, float(y) * 1000, float(z) * 1000]
                    edged_seq = np.vstack((edged_seq, preprocess_skeleton_frame(frame,
                                                                                is_15_joint=False,
                                                                                to_edge=True,
                                                                                to_rotate=self.is_rotated,
                                                                                is_25_joint=is_25_joint)))
                    seq = np.vstack((seq, preprocess_skeleton_frame(frame,
                                                                    is_15_joint=False,
                                                                    to_edge=False,
                                                                    to_rotate=self.is_rotated)))
            label_vector = len(labels) * np.ones(len(seq), dtype=np.int16)
            with open(label_file_path, 'r') as label_file:
                last_label_idx = 0
                for line in label_file:
                    if not line:
                        break   # EOF
                    if line[0].isalpha():
                        last_label_idx = labels.index(line.rstrip().lower())
                    else:
                        start, end = line.rstrip().split()
                        start, end = int(start) - offset, int(end) - offset
                        start -= self.forecast_t
                        if start < 0:
                            start = 0
                        label_vector[start:end+1] = last_label_idx
            if self.is_filtered:
                edged_seq = causal_savitzky_golay_filter(edged_seq)
                seq = causal_savitzky_golay_filter(seq)
            if self.is_edged and not self.is_translated:
                seq = edged_seq
            elif self.is_translated:
                seq = standardize_coordinate_origin_sequence(seq, is_25_joint=is_25_joint, use_hip=False)
                if self.is_edged:
                    seq = np.concatenate((seq, edged_seq), axis=-1)
            self._update_extrema(seq)
            self._data_label_pairs.append((seq, label_vector))
            pr.update(data_idx + 1)

    def load_data_by_idx(self, data_idx: int):
        return self._data_label_pairs[data_idx]

    # overrides
    def get_data_arrays(self, idx):
        return self.load_data_by_idx(idx)

    def __str__(self):
        return 'OAD_Dataset'


class SkeletonDatasetG3D(SkeletonDataset):
    """
    Load G3D Dataset by:
        V. Bloom, D. Makris, and V. Argyriou, "G3d: A gaming action dataset and real time action recognition evaluation
        framework," in Computer Vision and Pattern Recognition Workshops (CVPRW), 2012 IEEE Computer Society Conference
        on, 2012, pp. 7-12: IEEE.

    Dataset description:
        G3D dataset contains a range of gaming actions captured with Microsoft Kinect. The Kinect enabled us to record
        synchronised video, depth and skeleton data. The dataset contains 10 subjects performing 20 gaming actions:
        punch right, punch left, kick right, kick left, defend, golf swing, tennis swing forehand, tennis swing
        backhand, tennis serve, throw bowling ball, aim and fire gun, walk, run, jump, climb, crouch, steer a car,
        wave, flap and clap. The 20 gaming actions are recorded in 7 action sequences. Most sequences contain
        multiple actions in a controlled indoor environment with a fixed camera, a typical setup for gesture based
        gaming. Each sequence is repeated three times by each subject as shown in Table below.


        Actor	Fighting	Golf	    Tennis	    Bowling	    FPS	        Driving	    Misc
        1	    22	23	24	25	26	27	28	29	30	31	32	33*	34	35	36	37	38	39	40	41	42

        2	    43	44	45	46	47	48	49	50	51	52	53	54	55	56	57	58	59	60	61	62	63

        3	    64	65	66	67	68	69	70	71	72	73	74	75	76	77	78	79	80	81	82	83	84

        4	    85	86	87	88	89	90	91	92	93	94	95	96	97	98	99	100	101	102	103	104	105

        5	    106	107	108	109	110	111	112	113	114	115	116	117	118	119	120	121	122	123	124	125	126

        6	    127	128	129	130	131	132	133	134	135	136	137	138	139	140	141	142	143	144	145	146	147

        7	    148	149	150	151	152	153	154	155	156	157	158	159	160	161	162	163	164	165	166	167	168

        8	    169	170	171	172	173	174	175	176	177	178	179	180	181	182	183	184	185	186	187	188	189

        9	    190	191	192	193	194	195	196	197	198	199	200	201	202	203	204	205	206	207	208	209	210

        10	    214	215	216	217	218	219	220	221	222	223	224	225	226	227	228	229	230	231	232	233	234

        *Please note sequence 33 was corrupted and is therefore not available.

    Formats:
        Due to the formats selected, it is possible to view all the recorded data and metadata without any special
        software tools. The three streams were recorded at 30fps in a mirrored view. The depth and colour images were
        stored as 640x480 PNG files and the skeleton data in XML files.
    """
    def __init__(self, directory: str,
                 train_portion: float = 0.5,
                 regression_sigma: float = 5,
                 is_translated: bool = True,
                 is_edged: bool = False,
                 is_rotated: bool = False,
                 is_filtered: bool = False,
                 *args,
                 **kwargs):
        super(SkeletonDatasetG3D, self).__init__(directory, regression_sigma, train_portion,
                                                 is_translated=is_translated,
                                                 is_edged=is_edged,
                                                 is_rotated=is_rotated,
                                                 is_filtered=is_filtered,
                                                 downsample_factor=1,
                                                 forecast_t=10)
        self._labels = ('PunchRight', 'PunchLeft', 'KickRight', 'KickLeft', 'Defend',
                        'GolfSwing', 'TennisSwingForehand', 'TennisSwingBackhand',
                        'TennisServe', 'ThrowBowlingBall', 'AimAndFireGun', 'Walk',
                        'Run', 'Jump', 'Climb', 'Crouch', 'SteerCentre', 'SteerRight',
                        'SteerLeft', 'Wave', 'Flap', 'Clap', '~')

    def load_label_map(self):
        import re
        from lxml import etree
        labels = ('PunchRight', 'PunchLeft', 'KickRight', 'KickLeft', 'Defend',
                  'GolfSwing', 'TennisSwingForehand', 'TennisSwingBackhand',
                  'TennisServe', 'ThrowBowlingBall', 'AimAndFireGun', 'Walk',
                  'Run', 'Jump', 'Climb', 'Crouch', 'SteerCentre', 'SteerRight',
                  'SteerLeft', 'Wave', 'Flap', 'Clap')
        train_ids = [22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 34, 35, 36, 37, 38, 39, 40, 41, 42,
                     43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60, 61, 62, 63,
                     64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 84,
                     85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96, 97, 98, 99, 100, 101, 102, 103, 104, 105,
                     106, 107, 108, 109, 110, 111, 112, 113, 114, 115, 116, 117, 118, 119, 120, 121, 122, 123,
                     124, 125, 126]
        self._joints_per_person = SensorJointNumber.KINECT_V1
        category_folders, _ = get_folders_and_files(self.root_dir)
        pr = ProgressBar(80, 209)
        count = 0
        for category in category_folders:
            category_dir = self.root_dir + '/' + category + '/'
            sample_folders, _ = get_folders_and_files(category_dir)
            for sample in sample_folders:
                sample_id = int(sample.replace('KinectOutput', ''))
                if sample_id in train_ids:
                    is_train = True
                else:
                    is_train = False
                data_dir = category_dir + sample + '/Skeleton/'
                _, data_filenames = get_folders_and_files(data_dir)
                data_file_ids = []
                for filename in data_filenames:
                    if filename.endswith('.xml'):
                        data_file_ids.append(int(re.findall(r'\d+', filename)[0]))
                    else:
                        continue
                data_file_ids = sorted(data_file_ids)
                seq = np.array([], dtype=np.float32).reshape(0, self._joints_per_person * 3)
                edged_seq = np.array([], dtype=np.float32).reshape(0, self._joints_per_person * 3)
                frame_id_dict = {}
                for data_file_id in data_file_ids:
                    frame = []
                    skeleton_tree = etree.parse(data_dir + 'Skeleton ' + str(data_file_id) + '.xml')
                    joints = skeleton_tree.xpath('//Joint')
                    if len(joints) < 20:
                        continue
                    for joint in joints[:20]:
                        coords = joint[0]
                        frame.append(float(coords[0].text) * 1000.)  # x in mm
                        frame.append(float(coords[1].text) * 1000.)  # y in mm
                        frame.append(float(coords[2].text) * 1000.)  # z in mm
                    seq = np.vstack((seq, preprocess_skeleton_frame(frame,
                                                                    is_15_joint=False,
                                                                    to_edge=False,
                                                                    to_rotate=self.is_rotated)))
                    edged_seq = np.vstack((edged_seq, preprocess_skeleton_frame(frame,
                                                                                is_15_joint=False,
                                                                                to_edge=True,
                                                                                to_rotate=self.is_rotated)))
                    frame_id_dict[data_file_id] = len(seq) - 1

                label_tree = etree.parse(self.root_dir + 'ActionPoints' + str(sample_id) + '.xml')
                label_vector = len(labels) * np.ones(len(seq), dtype=np.int16)
                action_points = label_tree.xpath('//ActionPoint')
                prev_label_id = len(labels)
                prev_frame_idx = 0
                for action_point in action_points:
                    label_id = labels.index(action_point[0].text)
                    start_frame_idx = frame_id_dict[int(action_point[1].text)]
                    label_vector[prev_frame_idx:start_frame_idx] = prev_label_id
                    prev_frame_idx = start_frame_idx - self.forecast_t
                    prev_label_id = label_id
                label_vector[prev_frame_idx:] = prev_label_id   # annotate the very last action of the sequence
                if self.is_filtered:
                    seq = causal_savitzky_golay_filter(seq)
                    edged_seq = causal_savitzky_golay_filter(edged_seq)
                if self.is_edged and not self.is_translated:
                    seq = edged_seq
                elif self.is_translated:
                    seq = standardize_coordinate_origin_sequence(seq, use_hip=False)
                    if self.is_edged:
                        seq = np.concatenate((seq, edged_seq), axis=-1)
                self._data_label_pairs.append((seq, label_vector))
                self._update_extrema(seq)
                if is_train:
                    self._train_indices.append(count)   # not the same as the 'train_ids' variable
                else:
                    self._test_indices.append(count)
                count += 1
                pr.update(count)

    def load_data_by_idx(self, data_idx: int):
        return self._data_label_pairs[data_idx]

    # overrides
    def get_data_arrays(self, idx):
        return self.load_data_by_idx(idx)

    def __str__(self):
        return 'G3D_Dataset'


class SkeletonDatasetUTKinect(SkeletonDataset):
    """
    Load UTKinect-Action3D Dataset (untrimmed) by:
        L. Xia, C.-C. Chen, and J. K. Aggarwal, "View invariant human action recognition using histograms of 3d
        joints," in Computer vision and pattern recognition workshops (CVPRW), 2012 IEEE computer society conference
        on, 2012, pp. 20-27: IEEE.

    Dataset description:
        The videos was captured using a single stationary Kinect with Kinect for Windows SDK Beta Version. There are
        10 action types: walk, sit down, stand up, pick up, carry, throw, push, pull, wave hands, clap hands. There
        are 10 subjects, Each subject performs each actions twice. Three channels were recorded: RGB, depth and
        skeleton joint locations. The three channel are synchronized. The framerate is 30f/s. Note we only recorded
        the frames when the skeleton was tracked, so the frame number of the files has jumps. The final frame rate is
        about 15f/sec. (There is around 2% of the frames where there are multiple skeleton info recorded with
        slightly different joint locations. This is not caused by a second person. You can chooce either one.)

        In each video, the subject performs the 10 actions in a concatenate fation, the label of the each action
        segment is given in actionLabel.txt. The dataset contains 4 parts:

            (a) and (b) are RGB-D

            (c) Sketetal joint Locations (.txt):
                Each row contains the data of one frame, the first number is frame number, the following numbers
                are the (x,y,z) locations of joint 1-20. The x, y, and z are the coordinates relative to the sensor
                array, in meters.

            (d) Labels of action sequence (4KB)

    """
    def __init__(self, directory: str, train_portion: float = 0.5, regression_sigma: float = 5):
        super(SkeletonDatasetUTKinect, self).__init__(directory, regression_sigma, train_portion)

    def load_label_map(self):
        self._joints_per_person = SensorJointNumber.KINECT_V1

        # As the data loading also involves label tensor (where label are final indices) generations, pre-sorting is
        # required.
        self._labels = sorted(('walk', 'sitDown', 'standUp', 'pickUp', 'carry', 'throw', 'push', 'pull',
                               'waveHands', 'clapHands'))
        file_id = ''
        with open(self.root_dir + 'actionLabel.txt', 'r') as label_file:
            labels_and_time = []
            for line_idx, line in enumerate(label_file, 0):
                line = line.rstrip().replace(':', '')
                if line.strip() == '':
                    break   # EOF
                if line[0] == 's' and line[1].isdigit():
                    if line_idx != 0:
                        # store the last sequence
                        seq = np.array([], dtype=np.float32).reshape(0, self._joints_per_person * 3)
                        frame_label_indices = []
                        visited_frame_indices = []  # for discarding duplicate frames
                        with open(self.root_dir + 'joints/joints_' + file_id + '.txt', 'r') as data_file:
                            for data_line in data_file:
                                if data_line.rstrip() == '':
                                    break  # EOF
                                temp_list = data_line.split()
                                current_frame_idx = int(temp_list[0])
                                if current_frame_idx not in visited_frame_indices:
                                    frame = []
                                    for joint_idx in range(self._joints_per_person):
                                        frame.append(float(temp_list[joint_idx * 3 + 1]) * 1000.)  # x
                                        frame.append(float(temp_list[joint_idx * 3 + 2]) * 1000.)  # y
                                        frame.append(float(temp_list[joint_idx * 3 + 3]) * 1000.)  # z
                                    seq = np.vstack((seq, np.array(frame, dtype=np.float32)))
                                    for idx, time_label_pair in enumerate(labels_and_time, 0):
                                        if current_frame_idx >= time_label_pair[1]:
                                            if current_frame_idx <= time_label_pair[2]:
                                                # frame belongs to an action class
                                                frame_label_indices.append(self._labels.index(time_label_pair[0]))
                                                break
                                            else:
                                                if idx != len(labels_and_time) - 1:
                                                    next_time_label_pair = labels_and_time[idx + 1]
                                                    if current_frame_idx < next_time_label_pair[1]:
                                                        # unknown action
                                                        frame_label_indices.append(self.label_size)
                                                        break
                                                else:
                                                    # unknown action
                                                    frame_label_indices.append(self.label_size)
                                                    break
                                        else:
                                            # unknown action at the very beginning of the sequence
                                            frame_label_indices.append(self.label_size)
                                            break
                                    visited_frame_indices.append(current_frame_idx)
                        self._data_label_pairs.append((seq, np.array(frame_label_indices, dtype=np.int16)))
                        self._update_extrema(seq)

                    # prepare to record the next sequence
                    file_id = line
                    labels_and_time = []
                    continue
                temp_list = line.split()
                label_name = temp_list[0]
                if temp_list[1] == 'NaN' or temp_list[2] == 'NaN':  # Invalid frame numbers
                    continue
                labels_and_time.append((label_name, int(temp_list[1]), int(temp_list[2])))

    def load_data_by_idx(self, data_idx: int):
        return self._data_label_pairs[data_idx]

    # overrides
    def get_data_arrays(self, idx):
        return self.load_data_by_idx(idx)

    def __str__(self):
        return 'UTKinect_Dataset'
